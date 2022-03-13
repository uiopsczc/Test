import sys
import os

sys.path.append(os.path.dirname(os.path.realpath(__file__ + "/..")))  # 一定要放到这里（值为py_tools文件夹），不然会import不到module的
from openpyxl import load_workbook
from export_xlsx.ExportXlsx2Cs import *
from export_xlsx.ExportXlsx2Json import *
from export_xlsx.ExportXlsx2Lua import *
from export_xlsx.ExportXlsxConst import *
from export_xlsx.ExportXlsxUtil import *
from export_xlsx.SheetCfg import *
from pythoncat.util.FileUtil import *


def FilterFilePath(filePath):
  if not filePath.endswith(".xlsx") or filePath.find("~$") != -1:  # "~$"临时打开的文件过滤掉
    return True
  return False


# Export所有的文件
def ResetAll():
  if ExportXlsxConst.Is_Export_Lua:
    ExportXlsx2Lua.ResetAll()
  if ExportXlsxConst.Is_Export_Cs:
    ExportXlsx2Cs.ResetAll()
  ExportXlsx2Json.ResetAll()



# 根据filePath获取需要导出到的文件路径
def GetExportRelativeDirPath(filePath):
  relativeDirPath = filePath.replace(ExportXlsxConst.Xlsx_Dir_Path, "")
  slashStartIndex = relativeDirPath.rfind("\\")
  if slashStartIndex == -1:
    result = ""
  else:
    result = relativeDirPath[0:slashStartIndex + 1]
  return result


# Export所有的文件
def ExportAll():
  filePathList = FileUtil.GetFilePathList(ExportXlsxConst.Xlsx_Dir_Path, FilterFilePath)
  for filePath in filePathList:
    exportRelativeFilePath = filePath.replace(ExportXlsxConst.Xlsx_Dir_Path, "")
    print("正在导出%s" % exportRelativeFilePath)
    exportRelativeDirPath = GetExportRelativeDirPath(filePath)
    workbook = load_workbook(filePath, read_only=True, data_only=True)
    ExportWorkbook(workbook, exportRelativeDirPath, exportRelativeFilePath)


def ExportWorkbook(workbook, exportRelativeDirPath, exportRelativeFilePath):
  sheetCount = len(workbook.sheetnames)
  for sheetIndex in range(0, sheetCount):
    sheet = workbook.worksheets[sheetIndex]
    ExportSheet(sheet, exportRelativeDirPath, exportRelativeFilePath)


def ExportSheet(sheet, exportRelativeDirPath, exportRelativeFilePath):
    print("  正在导出[%s]" % (exportRelativeDirPath + sheet.title))
    sheetCfgFieldInfoDict = SheetCfgFieldInfoDict(sheet)
    sheetCfg = SheetCfg(sheet, ExportXlsxConst.Name_Json, sheetCfgFieldInfoDict)
    if sheetCfg.IsOutput():
      jsonDict = ExportXlsx2Json.ExportSheet(sheet, sheetCfg)
    if ExportXlsxConst.Is_Export_Lua:
      sheetCfg = SheetCfg(sheet, ExportXlsxConst.Name_Lua, sheetCfgFieldInfoDict)
      if sheetCfg.IsOutput():
        ExportXlsx2Lua.ExportSheet(sheet, jsonDict, sheetCfg, exportRelativeFilePath)
    if ExportXlsxConst.Is_Export_Cs:
      sheetCfg = SheetCfg(sheet, ExportXlsxConst.Name_Cs, sheetCfgFieldInfoDict)
      if sheetCfg.IsOutput():
        ExportXlsx2Cs.ExportSheet(sheet, jsonDict, sheetCfg, exportRelativeFilePath)




def main():
  ResetAll()
  ExportAll()
  print("ExportXlsx finished")


main()
