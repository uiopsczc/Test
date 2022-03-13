import os

from openpyxl import Workbook, load_workbook

class ExcelUtil(object):
  # https://openpyxl.readthedocs.io/en/stable/
  @staticmethod
  def ReadExcelAsLineList(sheet, startRowIndex=1):# row_num base on 1
    maxRowIndex = sheet.max_row
    maxColumnIndex = sheet.max_column
    lineList = []
    for rowIndex in range(1, maxRowIndex + 1):  # 行
      if rowIndex < startRowIndex:
        continue
      line = []
      for columnIndex in range(1, maxColumnIndex + 1):  # 列
        value = sheet.cell(row=rowIndex, column=columnIndex).value
        line.append(value)
      if line[0] and line[0] != "":
        lineList.append(line)
    return lineList

  @staticmethod
  def ReadExcelAsLineListFromFilePath(filePath, startRowIndex=1, sheetIndex=0):  # row_num base on 1
    workbook = load_workbook(filePath, read_only=True, data_only=True)
    sheet = workbook.worksheets[sheetIndex]
    return ExcelUtil.ReadExcelAsLineList(sheet, startRowIndex)


  @staticmethod
  def ReadExcelAsLine(sheet, rowIndex=1):  # row_num base on 1
    maxColumnIndex = sheet.max_column
    maxRowIndex = sheet.max_row
    line = []
    if maxRowIndex < rowIndex:
      return line
    for columnIndex in range(1, maxColumnIndex + 1):  # 列
      value = sheet.cell(row=rowIndex, column=columnIndex).value
      line.append(value)
    return line

  @staticmethod
  def ReadExcelAsLineFromFilePath(filePath, rowIndex=1, sheetIndex=0):  # row_num base on 1
    workbook = load_workbook(filePath, read_only=True, data_only=True)
    sheet = workbook.worksheets[sheetIndex]
    return ExcelUtil.ReadExcelAsLine(sheet, rowIndex)


  @staticmethod
  def WriteExcelFromLineList(sheet, lineList, startRowIndex=1):
    rowIndex = startRowIndex
    for line in lineList:
      columnIndex = 1
      for field in line:
        sheet.cell(row=rowIndex, column=columnIndex).value = field
        columnIndex += 1
      rowIndex += 1

  @staticmethod
  def WriteExcelFromLineListFromFilePath(filePath, lineList, startRowIndex=1, sheetIndex=0):
    if not os.path.exists(filePath):
      workbook = Workbook()
    else:
      workbook = load_workbook(filePath, data_only=True)  # data_only,读取公式的结果，而不是公式本身
    sheet = workbook.worksheets[sheetIndex]
    ExcelUtil.WriteExcelFromLineList(sheet, lineList, startRowIndex)
    return workbook

  # 删除空的，没有数据的行
  @staticmethod
  def ClearExcelEmptyRows(sheet, startRowIndex=1):
    maxRowIndex = sheet.max_row
    maxColumnIndex = sheet.max_column
    for rowIndex in range(maxRowIndex + 1, 0, -1):  # 行
      if rowIndex < startRowIndex:
        continue
      isHasData = False
      for columnIndex in list(range(1, maxColumnIndex + 1)):  # 行
        if sheet.cell(rowIndex, columnIndex).value is not None:
          isHasData = True
          break
      if not isHasData:
        sheet.delete_rows(rowIndex)

  @staticmethod
  def ClearExcelEmptyRowsFromFilePath(filePath, startRowIndex=1, sheetIndex=0):
    workbook = load_workbook(filePath, data_only=True)  # data_only,读取公式的结果，而不是公式本身
    sheet = workbook.worksheets[sheetIndex]
    ExcelUtil.ClearExcelEmptyRows(sheet, startRowIndex)
    return workbook
